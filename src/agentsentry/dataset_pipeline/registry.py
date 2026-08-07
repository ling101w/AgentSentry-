from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from .io import normalize_text, sha256_file, stable_id
from .sources import resolve_source


REQUIRED_HEADERS = ("数据集", "威胁", "分类依据", "边界说明")
_THREAT_PATTERN = re.compile(r"(?<![A-Z0-9])T([1-6])(?!\d)", re.IGNORECASE)


class RegistryImportError(ValueError):
    """Raised when a source registry workbook does not match the contract."""


def import_registry_xlsx(path: str | Path, *, sheet: str | None = None) -> list[dict[str, Any]]:
    """Import the four-column dataset registry workbook without changing its content.

    Registry metadata is normalized with NFKC for reliable source/code matching.
    Security sample payloads use the separate lossless text path and are not
    processed here. The workbook hash and physical Excel row remain traceable.
    """

    source = Path(path)
    if not source.is_file():
        raise RegistryImportError(f"dataset registry workbook not found: {source}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised in an isolated env
        raise RegistryImportError(
            "Excel registry import requires openpyxl; install it with "
            "`python -m pip install openpyxl`."
        ) from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise RegistryImportError(f"cannot open dataset registry workbook {source}: {exc}") from exc

    try:
        if sheet is not None:
            if sheet not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise RegistryImportError(
                    f"worksheet {sheet!r} not found in {source}; available: {available}"
                )
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active

        row_iterator = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(row_iterator)
        except StopIteration as exc:
            raise RegistryImportError(f"dataset registry worksheet {worksheet.title!r} is empty") from exc

        headers = _normalized_columns(header_values)
        if headers != REQUIRED_HEADERS:
            expected = " / ".join(REQUIRED_HEADERS)
            actual = " / ".join(headers) if headers else "<empty>"
            raise RegistryImportError(
                f"invalid registry headers in {source}:{worksheet.title}!1; "
                f"expected exactly {expected}, got {actual}"
            )

        workbook_sha256 = sha256_file(source)
        rows: list[dict[str, Any]] = []
        seen_source_ids: dict[str, int] = {}
        for source_row, values in enumerate(row_iterator, start=2):
            normalized = tuple(_registry_text(value) for value in values)
            if not any(normalized):
                continue
            if len(normalized) > len(REQUIRED_HEADERS) and any(normalized[len(REQUIRED_HEADERS) :]):
                raise RegistryImportError(
                    f"{source}:{worksheet.title}!{source_row}: unexpected data after 边界说明"
                )
            padded = normalized[: len(REQUIRED_HEADERS)] + ("",) * max(
                0, len(REQUIRED_HEADERS) - len(normalized)
            )
            dataset_raw, threat_raw, classification_basis, boundary_raw = padded
            if not dataset_raw:
                raise RegistryImportError(
                    f"{source}:{worksheet.title}!{source_row}: 数据集 must not be empty"
                )
            threat_codes, threat_name = parse_threat(threat_raw, source_row=source_row)
            if not classification_basis:
                raise RegistryImportError(
                    f"{source}:{worksheet.title}!{source_row}: 分类依据 must not be empty"
                )

            # Resolve both forms: NFKC can turn ``Bench（MSB）`` into
            # ``Bench(MSB)``, while the declared alias intentionally retains
            # its original punctuation spacing.
            spec = resolve_source(dataset_raw) or resolve_source(str(values[0]).strip())
            dataset = spec.dataset if spec is not None else dataset_raw
            source_id = spec.key if spec is not None else stable_id(dataset, prefix="source")
            previous_row = seen_source_ids.get(source_id)
            if previous_row is not None:
                raise RegistryImportError(
                    f"{source}:{worksheet.title}!{source_row}: duplicate dataset {dataset_raw!r}; "
                    f"first seen at row {previous_row}"
                )
            seen_source_ids[source_id] = source_row

            secondary = list(threat_codes[1:])
            if spec is not None:
                secondary.extend(spec.threat_secondary)
            threat_primary = threat_codes[0]
            threat_secondary = sorted(
                {code for code in secondary if code != threat_primary}, key=lambda code: int(code[1:])
            )
            boundary_note = None if boundary_raw == "—" or not boundary_raw else boundary_raw
            rows.append(
                {
                    "source_id": source_id,
                    "dataset": dataset,
                    "dataset_raw": dataset_raw,
                    "threat_primary": threat_primary,
                    "threat_secondary": threat_secondary,
                    "threat_name": threat_name,
                    "classification_basis": classification_basis,
                    "boundary_note": boundary_note,
                    "repo_url": spec.repo_url if spec is not None else None,
                    "license": None,
                    "version": None,
                    "raw_format": spec.raw_format if spec is not None else None,
                    "sample_count": None,
                    "has_benign": None,
                    "has_attack": None,
                    "has_trajectory": None,
                    "has_tool_context": None,
                    "has_ground_truth": None,
                    "adapter": spec.key if spec is not None else None,
                    "agent_sentry_adapter": spec is not None,
                    "status": "TODO",
                    "notes": None,
                    "source_file": source.name,
                    "source_sha256": workbook_sha256,
                    "source_sheet": worksheet.title,
                    "source_row": source_row,
                }
            )
        return rows
    finally:
        workbook.close()


def import_registry(path: str | Path, *, sheet: str | None = None) -> list[dict[str, Any]]:
    """Compatibility alias for callers that do not need to name the file format."""

    return import_registry_xlsx(path, sheet=sheet)


def enrich_registry_rows(
    rows: Iterable[Mapping[str, Any]],
    records: Iterable[Mapping[str, Any]],
    manifest: Mapping[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Backfill collected/build status without changing the source workbook."""

    records_by_source: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for record in records:
        source = record.get("source")
        if isinstance(source, Mapping):
            dataset = normalize_text(source.get("dataset"))
            if dataset:
                records_by_source[dataset.casefold()].append(record)

    manifest_by_source: dict[str, Mapping[str, Any]] = {}
    manifest_rows = manifest.get("sources", []) if isinstance(manifest, Mapping) else []
    if isinstance(manifest_rows, list):
        for item in manifest_rows:
            if isinstance(item, Mapping) and normalize_text(item.get("dataset")):
                manifest_by_source[normalize_text(item.get("dataset")).casefold()] = item

    enriched: list[dict[str, Any]] = []
    for original in rows:
        row = dict(original)
        dataset_key = normalize_text(row.get("dataset")).casefold()
        source_records = records_by_source.get(dataset_key, [])
        source_manifest = manifest_by_source.get(dataset_key, {})
        if source_manifest:
            row["repo_url"] = source_manifest.get("url") or row.get("repo_url")
            row["license"] = source_manifest.get("license") or row.get("license")
            row["version"] = source_manifest.get("commit") or row.get("version")
            row["raw_format"] = source_manifest.get("raw_format") or row.get("raw_format")
            row["raw_sha256"] = source_manifest.get("sha256") or source_manifest.get("tree_sha256")
            row["downloaded_at"] = source_manifest.get("downloaded_at")
            row["status"] = "DOWNLOADED"
        if source_records:
            attack_values = [_label_attack(record) for record in source_records]
            row["sample_count"] = len(source_records)
            row["has_attack"] = any(value is True for value in attack_values)
            row["has_benign"] = any(value is False for value in attack_values)
            row["has_trajectory"] = any(_has_trajectory(record) for record in source_records)
            row["has_tool_context"] = any(_has_tool_context(record) for record in source_records)
            row["has_ground_truth"] = any(_has_ground_truth(record.get("raw")) for record in source_records)
            row["agent_sentry_adapter"] = True
            row["status"] = (
                "READY"
                if all(_quality_status(record) == "valid" for record in source_records)
                else "CLEANED"
            )
        enriched.append(row)
    return enriched


def parse_threat(value: Any, *, source_row: int | None = None) -> tuple[list[str], str]:
    """Parse one or more ordered T1-T6 labels and the human-readable threat name."""

    text = _registry_text(value)
    codes: list[str] = []
    for match in _THREAT_PATTERN.finditer(text):
        code = f"T{match.group(1)}"
        if code not in codes:
            codes.append(code)
    if not codes:
        location = f" at row {source_row}" if source_row is not None else ""
        raise RegistryImportError(
            f"invalid 威胁{location}: expected at least one code from T1 through T6, got {text!r}"
        )
    threat_name = _THREAT_PATTERN.sub(" ", text)
    threat_name = threat_name.strip(" \t/,，;；:：-—–|()[]")
    return codes, threat_name


def _normalized_columns(values: Iterable[Any]) -> tuple[str, ...]:
    columns = [_registry_text(value) for value in values]
    while columns and not columns[-1]:
        columns.pop()
    return tuple(columns)


def _registry_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", normalize_text(value))


def _label_attack(record: Mapping[str, Any]) -> bool | None:
    labels = record.get("labels")
    value = labels.get("attack") if isinstance(labels, Mapping) else None
    return value if isinstance(value, bool) else None


def _quality_status(record: Mapping[str, Any]) -> str:
    quality = record.get("quality")
    return normalize_text(quality.get("status")) if isinstance(quality, Mapping) else ""


def _has_trajectory(record: Mapping[str, Any]) -> bool:
    content = record.get("content")
    return bool(content.get("trajectory")) if isinstance(content, Mapping) else False


def _has_tool_context(record: Mapping[str, Any]) -> bool:
    content = record.get("content")
    agentsentry = record.get("agentsentry")
    return bool(
        isinstance(content, Mapping)
        and (normalize_text(content.get("tool_name")) or normalize_text(content.get("tool_response")))
    ) or bool(isinstance(agentsentry, Mapping) and normalize_text(agentsentry.get("tool")))


def _has_ground_truth(value: Any) -> bool:
    if isinstance(value, Mapping):
        for key, item in value.items():
            normalized_key = normalize_text(key).casefold().replace("_", " ")
            if normalized_key in {
                "ground truth",
                "ground truth output",
                "expected achievements",
                "expected result",
                "result",
                "evaluation criteria",
            } and item not in (None, "", [], {}):
                return True
            if normalized_key == "source" and isinstance(item, str) and "def ground_truth" in item:
                return True
            if _has_ground_truth(item):
                return True
    elif isinstance(value, list):
        return any(_has_ground_truth(item) for item in value)
    return False
